import sys
import os
from matplotlib import pyplot as plt
import numpy as np
import pandas as pd
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel,
                             QSizePolicy, QDateEdit, QRadioButton, QButtonGroup,
                             QFormLayout, QDialog, QListWidget, QListWidgetItem, 
                             QAbstractItemView, QLineEdit,QTabWidget, QTableWidget, QTableWidgetItem, QVBoxLayout,QTextEdit)
from PyQt5.QtCore import Qt, QDate, pyqtSignal
from PyQt5.QtGui import QFont
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.dates as mdates  
import pandas as pd
import calendar
import datetime
from datetime import date
import holidays
from scipy.stats import norm
from PyQt5.QtGui import QColor
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment,Border,Side
from PyQt5.QtWidgets import QGroupBox

script_directory = os.path.dirname(os.path.abspath(__file__))
excel_file_name = 'VacationRatex2.xlsx'
excel_file_path = os.path.join(script_directory, excel_file_name)
excel_file_path_new=os.path.join(script_directory, excel_file_name)
# Load sheets into DataFrames
df_absences = pd.read_excel(excel_file_path_new, sheet_name='Absences')
df_projects = pd.read_excel(excel_file_path_new, sheet_name='Projects')
df_managers = pd.read_excel(excel_file_path_new, sheet_name='Employee Entry')
df_managers.rename(columns={'Pers.No.': 'Employee ID'}, inplace=True)
df_managers.rename(columns={'Resource Manager': 'Manager Name'},inplace=True)
df_projects.rename(columns={'Engineer Name': 'Employee Name'}, inplace=True)
df_managers.rename(columns={'Personnel Number':'Employee Name'},inplace=True)
# Sort the DataFrame by Employee Name and Start Date
df_projects.sort_values(by=['Employee Name', 'Mission start date'], inplace=True)

# Initialize a list to store rows with intercontract periods
intercontract_rows = []

# Iterate over unique employees
for employee, employee_data in df_projects.groupby('Employee Name'):
    # Iterate over projects for each employee
    for i in range(len(employee_data) - 1):
        current_end_date = employee_data.iloc[i]['Mission end date']
        next_start_date = employee_data.iloc[i + 1]['Mission start date']
        # Check for gap between projects
        if current_end_date < next_start_date:
            # Insert intercontract period
            intercontract_rows.append({
                'Employee ID': employee_data.iloc[i]['Employee ID'],
                'Employee Name': employee,
                'Mission start date': current_end_date,
                'Mission end date': next_start_date,
                'Project Name': 'intercontract'
            })

# Create DataFrame for intercontract periods
intercontract_df = pd.DataFrame(intercontract_rows)
# Concatenate original DataFrame with intercontract DataFrame
df_projects = pd.concat([df_projects, intercontract_df]).sort_index().reset_index(drop=True)
employees_projects_only = df_projects[~df_projects['Employee ID'].isin(df_absences['Employee ID'])]
# Identify intercontract and newcomer employees
current_date = datetime.datetime.now()
two_months_ago = current_date - datetime.timedelta(days=60)
two_months_ago=pd.to_datetime(two_months_ago)
modified_df_manager=df_managers.copy()
modified_df_manager['Start Date'] = pd.to_datetime(df_managers['Start Date'])

intercontract_employees = df_managers[modified_df_manager['Start Date'] <= two_months_ago]
newcomer_employees = df_managers[modified_df_manager['Start Date'] > two_months_ago]
# Check if intercontract and newcomer employees are not already in the first two sheets
intercontract_employees = intercontract_employees[~intercontract_employees['Employee ID'].isin(df_projects['Employee ID'])]

newcomer_employees = newcomer_employees[~newcomer_employees['Employee ID'].isin(df_absences['Employee ID']) &
                                        ~newcomer_employees['Employee ID'].isin(df_projects['Employee ID'])]

intercontract_employees['Project Name'] = 'Intercontract'
newcomer_employees['Project Name'] = 'Newcomer'

intercontract_employees_df = pd.DataFrame({
    'Employee ID': intercontract_employees['Employee ID'],
    'Employee Name': intercontract_employees['Employee Name'],
    'Project Name': intercontract_employees['Project Name'],  # Include project information
    'End Customer': intercontract_employees['Project Name'],  # Include End Customer information
    'Att./abs. days': '0',
    'Calendar days': '0',
    'Request Status': 'Rejected',
    'Sum of Entitlement': '25',
    'Absence Type': 'Annual leave',
    'From': df_absences['From'].min(),
    'To': datetime.datetime.now(),
    'Request Status': 'Rejected'  # Add 'Request Status' column with value 'Rejected'
})

newcomer_employees_df = pd.DataFrame({
    'Employee ID': newcomer_employees['Employee ID'],
    'Employee Name': newcomer_employees['Employee Name'],
    'Project Name': newcomer_employees['Project Name'],  # Include project information
    'End Customer': newcomer_employees['Project Name'],  # Include End Customer information
    'Att./abs. days': '0',
    'Calendar days': '0',
    'Request Status': 'Rejected',
    'Sum of Entitlement': '25',
    'Absence Type': 'Annual leave',
    'From': newcomer_employees['Start Date'],
    'To': datetime.datetime.now(),
    'Request Status': 'Rejected'  # Add 'Request Status' column with value 'Rejected'
})
# Add intercontract and newcomer employees to merged_df

missing_employees_df = pd.DataFrame({
    'Employee ID': employees_projects_only['Employee ID'],
    'Employee Name': employees_projects_only['Employee Name'],
    'Project Name': employees_projects_only['Project Name'],  # Include project information
    'End Customer': employees_projects_only['End Customer'],  # Include End Customer information
    'Att./abs. days': 0,
    'Calendar days': 0,
    'Sum of Entitlement': 25,
    'Absence Type': 'Annual leave',
    'From': df_absences['From'].min(),
    'To': df_absences['To'].max(),
   })


# Merge managers' information with final_df based on Employee ID

# Save the modified DataFrame to Excel

df_projects.rename(columns={'Employee Name': 'Engineer Name'}, inplace=True)
merged_df = pd.merge(df_absences, df_projects.drop(columns=['Engineer Name']), on='Employee ID', how='inner')
# Filter merged DataFrame based on condition
filtered_df = merged_df[(merged_df['From'] >= merged_df['Mission start date']) & (merged_df['From'] <= merged_df['Mission end date'])]
filtered_df= pd.concat([filtered_df, missing_employees_df,newcomer_employees_df,intercontract_employees_df], ignore_index=True)
# Find employees present in Absences but not in Projects
employees_absences_only = df_absences[~df_absences['Employee ID'].isin(df_projects['Employee ID'])]     

# Create a DataFrame for these employees with 'No contract' as End Customer and Project Name
employees_absences_only['End Customer'] = 'No contract'
employees_absences_only['Project Name'] = 'No contract'
# Save modified DataFrame back to the 'Absences' sheet
# Find employees who took days off without a project assigned

employees_wrong_dates = pd.merge(df_absences, df_projects, on='Employee ID', how='inner')
employees_wrong_dates = employees_wrong_dates[~((employees_wrong_dates['From'] >= employees_wrong_dates['Mission start date']) & (employees_wrong_dates['From'] <= employees_wrong_dates['Mission end date']))]


# Add 'No project assigned' as Project Name for these employees
employees_wrong_dates['End Customer'] = 'Intercontract'
employees_wrong_dates['Project Name'] = 'Intercontract'

final_df = pd.concat([filtered_df, employees_absences_only, employees_wrong_dates, missing_employees_df], ignore_index=True)
final_df = pd.merge(final_df, df_managers, on='Employee ID', how='left')
final_df.drop(columns=['Engineer Name'], inplace=True)
if 'Manager Name_x' in final_df.columns and 'Manager Name_y' in final_df.columns:
    final_df['Manager Name_x'].fillna(final_df['Manager Name_y'], inplace=True)
    # Drop the redundant 'Manager Name_y' column
    final_df.drop(columns=['Manager Name_y'], inplace=True)
final_df.rename(columns={'Manager Name_x': 'Manager Name'}, inplace=True)
final_df.rename(columns={'Employee Name_x':'Employee Name'},inplace=True)
final_df.drop(columns=['Employee Name_y'], inplace=True)
final_df.to_excel('vacationRate_modified.xlsx', index=False, sheet_name='Absences')

excel_final_name ='vacationRate_modified.xlsx'
excel_final_path = os.path.join(script_directory, excel_final_name)
df = pd.read_excel(excel_final_path)

unique_managers=df["Manager Name"].unique().tolist()
unique_departments = df["Departament"].unique().tolist()
unique_project = df["Project Name"].unique().tolist()
unique_employee = df["Employee Name"].unique().tolist()
unique_leave = df["Absence Type"].unique().tolist()
# Define a modern-looking stylesheet for the application
stylesheet = """
    QMainWindow {
        background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(173, 216, 230, 255), stop:1 rgba(255, 255, 255, 255));
    }
    QPushButton {
        font-size: 12px; /* Adjust size as needed */
        font-weight: bold; /* Make text bold */
        background-color: rgb(52, 154, 255);  /* Changed button color here */
        border-radius: 5px;
        color: white;
        padding: 6px;
        margin: 6px;
        font-size: 16px;
    }
    QPushButton:hover {
        background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 180, 255), stop:1 rgba(0, 0, 230, 255));
    }
    QPushButton:pressed {
        background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 100, 255), stop:1 rgba(0, 0, 150, 255));
    }
    QLabel, QRadioButton {
        color: #003366;
    }
    QGroupBox {
        border: 2px solid #000080; /* Dark blue border */
        border-radius: 5px;
        margin-top: 7ex; /* Leave space at the top for the title */
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 rgba(240, 248, 255, 0.7),
                                    stop:1 rgba(200, 220, 255, 0.7)); /* Gradient background */
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        subcontrol-position: top center;
        padding: 0 3px;
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 rgba(240, 248, 255, 0.7),
                                    stop:1 rgba(200, 220, 255, 0.7)); /* Gradient background */
        color: #003366;
        /* Make the title bigger and bold */
        font-size: 18px; /* Adjust the font size as needed */
        font-weight: bold; /* Make the font bold */
        padding-top: 20px;
        padding: 3px 10px; /* Adjust the padding to make the title box bigger */
    }
    QGroupBox:hover {
        border: 2px solid #4169E1; /* Brighter blue border on hover */
    }

"""

# Custom Dialog for Period Selection
# This class allows users to select a reporting period for data visualization


class PeriodDialog(QDialog):
    
    customPeriodSelected = pyqtSignal(str, str)  # Signal for custom period (start date, end date)
    predefinedPeriodSelected = pyqtSignal(str)  # Signal for predefined period

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Period")
        self.setupUI()

    def setupUI(self):
        # Layout setup, including date pickers and period selection radio buttons
        layout = QVBoxLayout(self)

        # Date pickers for custom period
        formLayout = QFormLayout()
        self.startDateEdit = QDateEdit(self)
        self.startDateEdit.setDate(QDate.currentDate())
        self.endDateEdit = QDateEdit(self)
        self.endDateEdit.setDate(QDate.currentDate())
        formLayout.addRow('Start Date:', self.startDateEdit)
        formLayout.addRow('End Date:', self.endDateEdit)

        # Radio buttons for predefined periods
        self.radioButtons = []
        self.radioGroup = QButtonGroup(self)
        self.radioCustom = QRadioButton("Custom")
        self.radioCustom.setChecked(True)
        self.radioGroup.addButton(self.radioCustom)
        formLayout.addRow(self.radioCustom)

        # Add radio buttons for each month
        for i in range(1, 13):
            month_radio = QRadioButton(QDate.longMonthName(i))
            self.radioGroup.addButton(month_radio)
            self.radioButtons.append(month_radio)
            formLayout.addRow(month_radio)

        # Connect radio button group
        self.radioGroup.buttonClicked.connect(self.onRadioButtonClicked)

        # Add radio button for "All Available Periods"
        self.radioAllPeriods = QRadioButton("All Available Period")
        self.radioGroup.addButton(self.radioAllPeriods)
        formLayout.addRow(self.radioAllPeriods)


        # Submit and Cancel buttons
        buttonsLayout = QHBoxLayout()
        self.submitButton = QPushButton('Submit', self)
        self.cancelButton = QPushButton('Cancel', self)
        buttonsLayout.addWidget(self.submitButton)
        buttonsLayout.addWidget(self.cancelButton)

        # Connect buttons
        self.submitButton.clicked.connect(self.onSubmit)
        self.cancelButton.clicked.connect(self.close)

        # Set the dialog layout
        layout.addLayout(formLayout)
        layout.addLayout(buttonsLayout)
        self.setLayout(layout)  # Set the layout manager

        # Set cursor for buttons
        self.setCursorForButtons()

    def setCursorForButtons(self):
        for button in self.findChildren(QPushButton):
            button.setCursor(Qt.PointingHandCursor)

    def onRadioButtonClicked(self, button):
        if button == self.radioCustom:
            self.startDateEdit.setEnabled(True)
            self.endDateEdit.setEnabled(True)
        else:
            self.startDateEdit.setEnabled(False)
            self.endDateEdit.setEnabled(False)

    def onSubmit(self):
         # Handles submission, differentiating between custom and predefined periods
        checkedButton = self.radioGroup.checkedButton()
        if checkedButton:
            if checkedButton == self.radioCustom:
                startDate = self.startDateEdit.date().toString(Qt.ISODate)
                endDate = self.endDateEdit.date().toString(Qt.ISODate)
                print(f"Custom period: {startDate} to {endDate}")
                self.customPeriodSelected.emit(startDate, endDate)  # Emit the custom period signal
            else:
                month_name = checkedButton.text()
                print(f"Predefined month selected: {month_name}")
                self.predefinedPeriodSelected.emit(month_name)
        else:
            print("No period selection made.")
        self.close()

# Custom Dialog for Selection
class SelectionDialog(QDialog):
    selectionMade = pyqtSignal(list, str)
    def __init__(self, options, title, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.options = options
        self.setupUI()

    def setupUI(self):
        layout = QVBoxLayout(self)

        # Search input field
        self.searchInput = QLineEdit(self)
        self.searchInput.setPlaceholderText("Search ")
        self.searchInput.textChanged.connect(self.filterOptions)

        layout.addWidget(self.searchInput)
        self.setLayout(layout)

        # List widget for options
        self.listWidget = QListWidget(self)
        for option in self.options:
            item = QListWidgetItem(option)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            self.listWidget.addItem(item)
        self.listWidget.setSelectionMode(QAbstractItemView.MultiSelection)
        layout.addWidget(self.listWidget)

        # Select All, Submit, and Cancel buttons
        buttonsLayout = QHBoxLayout()
        self.selectAllButton = QPushButton('Select All', self)
        self.submitButton = QPushButton('Submit', self)
        self.cancelButton = QPushButton('Cancel', self)
        buttonsLayout.addWidget(self.selectAllButton)
        buttonsLayout.addWidget(self.submitButton)
        buttonsLayout.addWidget(self.cancelButton)

        # Connect buttons
        self.selectAllButton.clicked.connect(self.selectAll)
        self.submitButton.clicked.connect(self.onSubmit)
        self.cancelButton.clicked.connect(self.close)

        # Set the dialog layout
        layout.addLayout(buttonsLayout)

        # Set cursor for buttons
        self.setCursorForButtons()

    def setCursorForButtons(self):
        for button in self.findChildren(QPushButton):
            button.setCursor(Qt.PointingHandCursor)

    def filterOptions(self):
        search_text = self.searchInput.text().strip()
        if not search_text:
            # If the search input is empty, show all options
            for index in range(self.listWidget.count()):
                item = self.listWidget.item(index)
                item.setHidden(False)
        else:
            # Filter options based on search text
            for index in range(self.listWidget.count()):
                item = self.listWidget.item(index)
                item_text = item.text()
                item.setHidden(search_text.lower() not in item_text.lower())

    def selectAll(self):
        for index in range(self.listWidget.count()):
            item = self.listWidget.item(index)
            item.setCheckState(Qt.Checked)

    def onSubmit(self):
        selectedOptions = [self.listWidget.item(i).text() for i in range(self.listWidget.count()) 
                           if self.listWidget.item(i).checkState() == Qt.Checked]
        print(f"Selected options: {selectedOptions}")
        self.selectionMade.emit(selectedOptions, self.windowTitle())  # Emit the signal with selected options and the dialog title as category
        self.close()

class MonthlyTableWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Monthly Data")
        self.current_month = date.today().month
        self.current_year = date.today().year  # Assuming the initial year
        self.tableWidget = self.createMonthlyTable()
        layout = QVBoxLayout(self)
        self.resize(1600, 600)  # Adjust the width and height as needed
        self.setLayout(layout)

        previous_month = QDate(self.current_year, self.current_month, 1).addMonths(-1).toString("MMMM")
        next_month = QDate(self.current_year, self.current_month, 1).addMonths(1).toString("MMMM")

        # Set the text of the buttons
        self.currentMonthLabel = QLabel()
        current_month = QDate(self.current_year, self.current_month, 1).toString("MMMM")
        self.currentMonthLabel.setText(current_month)
        self.prevButton = QPushButton(previous_month)
        self.nextButton = QPushButton(next_month)
        self.exportButton = QPushButton("Export to Excel")  # Add export button
        self.exportButton.setSizePolicy(self.prevButton.sizePolicy())
        self.exportButton.setMaximumWidth(150)  # Set maximum width
        buttonLayout = QHBoxLayout()
        font = QFont()
        font.setFamily("Arial")
        font.setPointSize(20)  # Adjust the size as needed
        font.setBold(True)     # Make the font bold
        # Apply font to the label
        self.currentMonthLabel.setFont(font)
        self.currentMonthLabel.setAlignment(Qt.AlignCenter)
        buttonLayout.addWidget(self.prevButton)
        buttonLayout.addWidget(self.currentMonthLabel)
        buttonLayout.addWidget(self.nextButton)

        layout.addLayout(buttonLayout)
        layout.addWidget(self.tableWidget)
        
        # add a legend to the table
        self.tableLegendText = "S = Sick Leave \tM = Maternity Leave \nH = Annual Leave \tW = Wedding Leave \nU = Unpaid Leave \tF = Floating Day \nX = Unkown Type"
        self.tableLegend = QLabel()
        fontTableLegend = QFont()
        font.setFamily("Arial")
        font.setPointSize(8)
        self.tableLegend.setFont(fontTableLegend)
        self.tableLegend.setText(self.tableLegendText)

        layout.addWidget(self.tableLegend)

        layout.addLayout(buttonLayout)  # Add button layout under the table
        layout.addWidget(self.tableWidget)  # Add table widget
        layout.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.exportButton)  # Add export button below the table
        
        # Connect button clicks to slots
        self.prevButton.clicked.connect(self.showPreviousMonth)
        self.nextButton.clicked.connect(self.showNextMonth)
        self.exportButton.clicked.connect(self.export_to_excel)  # Connect export button to export function

    def getSelection(self):
        return ex.selections
    def get_monthly_data(self, year, month):
        filtered_data= self.getSelection()
        # Filter the DataFrame for the specified year and month
        monthly_data = df[(df['From'].dt.year == year) & ((df['From'].dt.month == month) | (df['To'].dt.month==month))]
        if filtered_data['employee'] is not None:
            monthly_data=monthly_data[monthly_data['Employee Name'].isin(filtered_data['employee'])]
            monthly_data=monthly_data.drop_duplicates(subset=['Employee Name', 'From'])
        else: 
            if filtered_data['project'] is not None:
                monthly_data=monthly_data[monthly_data['Project Name'].isin(filtered_data['project'])]
                
            else:
                if filtered_data['department'] is not None:
                    monthly_data = monthly_data[monthly_data['Departament'].isin(filtered_data['department'])]
                    monthly_data=monthly_data.drop_duplicates(subset=['Employee Name', 'From'])
                else:
                    if filtered_data['manager'] is not None:
                        monthly_data = monthly_data[monthly_data['Manager Name'].isin(filtered_data['manager'])]
                        monthly_data=monthly_data.drop_duplicates(subset=['Employee Name', 'From'])
        monthly_data=monthly_data[monthly_data['Request Status'].isin(['Approved','Requested'])]
        if not monthly_data.empty: 
            month_data = {}
            absence_list = []
            absence_type_list = []  # List to store absence types
            # Iterate over each row in the monthly data
            for index, row in monthly_data.iterrows():
                # Extract relevant information from the row
                absence_days=row['From']-row['To']
                
                employee_name = row['Employee Name']
                from_date = row['From'].day  # Extract day from 'From' column
                if row["From"].month<month:
                    row['From']=pd.Timestamp(year=self.current_year,month=month,day=1)
                    from_date=row['From'].day
                absence_days=row['To'].day-row['From'].day+1

                if row["To"].month>month:
                    absence_days=calendar.monthrange(self.current_year,month)[1]-from_date+1
                absence_type = row['Absence Type']  # Extract absence type from specified collumn
                if row["Att./abs. days"]==0.5:
                    absence_days=1.5
                absence_list.append(absence_days)
                absence_type_list.append(absence_type)
                # If the day is not already in the month_data dictionary, initialize it
                if from_date not in month_data:
                    month_data[from_date] = {}

                # Update the dictionary with employee absence days and absence type for the corresponding day
                month_data[from_date][employee_name] = (absence_days, absence_type)

            return month_data, absence_list, absence_type_list
        else:
            # If no data is found for the specified month and year, return empty dictionaries
            return {}, [], []
    def updateTableForMonth(self, month):
        # Clear the table
        self.tableWidget.clearContents()
        num_days = calendar.monthrange(self.current_year, month)[1]
        # Populate the table with data for the current month and year
        # Your existing code to populate the tableWidget for the current month
        self.updateTable(month)

    def export_to_excel(self):
        excel_final_name ='table.xlsx'
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), excel_final_name)

        # Create the workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        workbook.remove(worksheet)
        # Iterate over each month in the year
        for month in range(1, 13):
            # Generate the table for the current month
            self.updateTableForMonth(month)

            # Write data to the Excel worksheet
            self.writeDataToExcel(workbook, month)

            # Clear the tableWidget for the next month
            self.tableWidget.clearContents()

        # Save the workbook
        workbook.save(file_path)
        # Close the workbook
        workbook.close()
        self.updateTable(self.current_month)
        os.startfile(file_path)
    def writeDataToExcel(self, workbook, month):
        # Create a new worksheet for the month
        worksheet = workbook.create_sheet(title=calendar.month_name[month])
    
        # Write column names
        for col in range(1, self.tableWidget.columnCount()+1):
            header_item = self.tableWidget.horizontalHeaderItem(col-2)
            if header_item is not None:
                worksheet.cell(row=1, column=col, value=header_item.text())
    
        # Write row names
        for row in range(1, self.tableWidget.rowCount() + 1):
            header_item = self.tableWidget.verticalHeaderItem(row - 1)
            if header_item is not None:
                worksheet.cell(row=row + 1, column=1, value=header_item.text())
    
        # Write data along with background colors
        for row in range(self.tableWidget.rowCount()):
            for col in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(row, col)
                if item is not None:
                    cell = worksheet.cell(row=row + 2, column=col + 2)
                    cell.value = item.text()
    
                    # Convert Qt color to aRGB hex format
                    qt_color = item.background().color().name()
                    argb_hex_color = f"FF{qt_color[1:]}"  # Add alpha channel FF for full opacity
                    fill = PatternFill(start_color=argb_hex_color, end_color=argb_hex_color, fill_type="darkGrid")
                    cell.fill = fill
                    if col==0 or col == self.tableWidget.columnCount() - 1:
                        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="darkGrid")
                        cell.fill=fill
                # Set column widths based on the widths of the Qt table columns
        for col in range(self.tableWidget.columnCount()):
            qt_column_width = self.tableWidget.columnWidth(col)
            excel_column_letter = openpyxl.utils.get_column_letter(col + 2)  # Adjust for 1-based indexing in Excel
            worksheet.column_dimensions[excel_column_letter].width = qt_column_width / 7  # Convert from pixels to Excel units
        # Set alignment to wrap text for all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
    
        for col in range(self.tableWidget.columnCount()):
            qt_column_width = self.tableWidget.columnWidth(col)
            excel_column_letter = openpyxl.utils.get_column_letter(col + 2)  # Adjust for 1-based indexing in Excel
            worksheet.column_dimensions[excel_column_letter].width = qt_column_width / 7  # Convert from pixels to Excel units
    
        # Set alignment to wrap text for all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
        thinBorder=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border=thinBorder

    def updateTable(self,month):
        # Clear the table
        self.tableWidget.clearContents()
        num_days = calendar.monthrange(self.current_year, month)[1]
        self.tableWidget.setColumnCount(num_days + 2)
        light_blue = QColor(173, 216, 230)

        # Color the first two rows into light blue

        # Populate the table with data for the current month and year
        month_data, absence_list, absence_type_list = self.get_monthly_data(self.current_year, month)
        month_name = QDate.longMonthName(month).capitalize()
        self.tableWidget.setVerticalHeaderItem(0, QTableWidgetItem(month_name))
        for i in range(1,50):
         self.tableWidget.setVerticalHeaderItem(i, QTableWidgetItem(str(i)))   
        
        for i in range(1, num_days + 1):
            date = datetime.date(self.current_year, month, i)
            week_number = date.isocalendar()[1]  # Get the ISO week number
            week_str = f"CW{week_number:02d}"  # Format the week number
            self.tableWidget.setHorizontalHeaderItem(i, QTableWidgetItem(week_str))
        self.tableWidget.setHorizontalHeaderItem(0, QTableWidgetItem("Employee"))
        for i in range(1, num_days + 1):
            date = datetime.date(self.current_year, month, i)
            day_number = str(i)
            day_name = date.strftime("%A")  # Get the full name of the day
            header_text = f"{day_number}\n{day_name}"
            header_item = QTableWidgetItem(header_text)
            self.tableWidget.setItem(0, i, header_item)
            self.tableWidget.resizeRowsToContents()        
        self.tableWidget.setItem(0,num_days+1,QTableWidgetItem('Total'))
        self.tableWidget.resizeRowsToContents()
        
        seen_keys = set()
        ordered_names = []

        for inner_dict in month_data.values():
            for key in inner_dict.keys():
                if key not in seen_keys:
                    ordered_names.append(key)
                    seen_keys.add(key)
        for row_index, employee_name in enumerate(ordered_names, start=1):
            self.tableWidget.setItem(row_index, 0, QTableWidgetItem(employee_name.strip()))
            for day, inner_dict in month_data.items():
                absence_days = inner_dict.get(employee_name, [0])[0]  # Get absence days for the employee
                actual_day=datetime.date(self.current_year,month,day)
                pandas_day=pd.to_datetime(actual_day)

        # Update the table with the retrieved data
        for day, inner_dict in month_data.items():
            
            for employee_name, (absence_days, absence_type) in inner_dict.items():
              
                employee_name = employee_name.strip()  # Trim whitespace from employee name
                # Find the row index corresponding to the employee name
                items = self.tableWidget.findItems(employee_name, Qt.MatchExactly)
                if items:
                    row_index = items[0].row()
                    # Calculate the column index based on the day
                    column_index = day
                    # Set the absence days in the table cell
                    for i in range(int(absence_days)):
                        if column_index + i <= num_days:  # Ensure it doesn't exceed the maximum day
                            actual_day=datetime.date(self.current_year,month,column_index+i)
                            pandas_day=pd.to_datetime(actual_day)
                            if pandas_day.dayofweek<5:
                                color, _ = self.get_absence_type_color(absence_type)
                                cell_item = QTableWidgetItem(self.get_absence_letter(absence_type))
                                cell_item.setBackground(color)
                                self.tableWidget.setItem(row_index, column_index+i, cell_item)
                    
        
                            else:
                                cell_item = QTableWidgetItem('W')
                                light_grey = QColor(211,211,211)  # Adjust the RGB values for the desired shade of gray
                                cell_item.setBackground(light_grey)
                                self.tableWidget.setItem(row_index, column_index+i, cell_item)
                            if absence_days == 1.5:
                               cell_item = QTableWidgetItem('H1')
                               light_green = QColor(144, 238, 144)
                               cell_item.setBackground(light_green)
                               self.tableWidget.setItem(row_index, column_index+i, cell_item)
        for row_index in range(1, self.tableWidget.rowCount()):  # Iterate through rows
            total_days = 0.0
            for col_index in range(1, num_days + 1):  # Iterate through columns (days)
                item = self.tableWidget.item(row_index, col_index)
                if item is not None:
                    absence_days = item.text()  # Get absence days for the current cell
                    for day in absence_days:
                        print(day)
                        if  day == '1':
                            total_days -= 0.5
                        else:  
                            if day == 'H':
                                total_days += 1  
    
            # Add the total days taken to a new column at the end of the row
            total_days_item = QTableWidgetItem(str(total_days))
            self.tableWidget.setItem(row_index, num_days + 1, total_days_item)
        ro_holidays = self.get_national_holidays(self.current_year,month)

        light_grey = QColor(211,211,211)  # Adjust the RGB values for the desired shade of gray

        for row_index, employee_name in enumerate(ordered_names, start=1):
            for i in range(1, num_days + 1):
                if QDate(self.current_year, month, i) in ro_holidays:
                    cell_item = QTableWidgetItem('B')
                    cell_item.setBackground(light_grey)
                    self.tableWidget.setItem(row_index, i, cell_item)

        for col in range(self.tableWidget.columnCount()):
            item = self.tableWidget.item(0, col)
            if item is None:
                item = QTableWidgetItem()
                self.tableWidget.setItem(0, col, item)
            item.setBackground(light_blue)
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.repaint()
    def get_absence_type_color(self, absence_type):
        # Define colors for each absence type and their lighter versions
        color_map = {
            'Sick leave': (QColor(255, 192, 203), QColor(255, 224, 230)),
            'Maternity leave': (QColor(255, 165, 0), QColor(255, 204, 102)),
            'Annual leave': (QColor(104, 198, 104), QColor(255, 255, 153)),
            'Wedding leave': (QColor(0, 191, 255), QColor(153, 204, 255)),
            'Unpaid leave': (QColor(220, 20, 60), QColor(255, 179, 179)),
            'Floating day': (QColor(65, 105, 225), QColor(173, 216, 230)),
            'Bereavement leave': (QColor(169,169,169),QColor(211,211,211))
        }
        # Return the corresponding color, or default to white if absence type is not found
        return color_map.get(absence_type, (Qt.white, Qt.white))
    def get_absence_letter(self,absence_type):
        if absence_type == "Sick leave":
            return "S"
        elif absence_type == "Maternity leave":
            return "M"
        elif absence_type == "Annual leave":
            return "H"
        elif absence_type == "Wedding leave":
            return "W"
        elif absence_type == "Unpaid leave":
            return "U"
        elif absence_type == "Floating day":
            return "F"
        elif absence_type == "Bereavement leave":
            return "D"
        else:
            return "X"  # Default to 'X' for unknown absence types
    def get_national_holidays(self, year, month):
        # Create a Holidays object for Romania
        ro_holidays = holidays.RO(years=year)

        holidays_in_month = [date for date in ro_holidays.keys() if date.year == year and date.month == month]
        return holidays_in_month
    def showPreviousMonth(self):

        self.current_month -= 1
        if self.current_month < 1:
            self.current_month = 12
            self.current_year -= 1
        # Modify the previous and next buttons with correct months
        current_month = QDate(self.current_year, self.current_month, 1).toString("MMMM")            
        self.currentMonthLabel.setText(current_month)
        previous_month = QDate(self.current_year, self.current_month, 1).addMonths(-1).toString("MMMM")
        self.prevButton.setText(previous_month)
        next_month = QDate(self.current_year, self.current_month, 1).addMonths(+1).toString("MMMM")
        self.nextButton.setText(next_month)
        # Update the table with data for the next month
        self.updateTable(self.current_month)
        
    def showNextMonth(self):
        self.current_month += 1
        if self.current_month > 12:
            self.current_month = 1
            self.current_year += 1
        # Modify the previous and next buttons with correct months
        previous_month = QDate(self.current_year, self.current_month, 1).addMonths(-1).toString("MMMM")
        self.prevButton.setText(previous_month)
        current_month = QDate(self.current_year, self.current_month, 1).toString("MMMM")            
        self.currentMonthLabel.setText(current_month)
        next_month = QDate(self.current_year, self.current_month, 1).addMonths(1).toString("MMMM")
        self.nextButton.setText(next_month)
        # Update the table with data for the next month
        self.updateTable(self.current_month)

    def createMonthlyTable(self):
        tableWidget = QTableWidget()
        num_days = calendar.monthrange(self.current_year, self.current_month)[1]
        tableWidget.setColumnCount(num_days + 1)
        tableWidget.setRowCount(50)
        headers = [str(day) for day in range(1, 31)]  # Assuming maximum 31 days in a month

        tableWidget.setHorizontalHeaderLabels(headers)
        light_blue = QColor(173, 216, 230)

        # Color the first two rows into light blue

        # Populate the table with data for the current month and year
        month_data, absence_list, absence_type_list = self.get_monthly_data(self.current_year, self.current_month)
        month_name = QDate.longMonthName(self.current_month).capitalize()
        tableWidget.setVerticalHeaderItem(0, QTableWidgetItem(month_name))
        for i in range(1,50):
            tableWidget.setVerticalHeaderItem(i, QTableWidgetItem(str(i)))   
        
        for i in range(1, num_days + 1):
            date = datetime.date(self.current_year, self.current_month, i)
            week_number = date.isocalendar()[1]  # Get the ISO week number
            week_str = f"CW{week_number:02d}"  # Format the week number
            tableWidget.setHorizontalHeaderItem(i, QTableWidgetItem(week_str))
        tableWidget.setHorizontalHeaderItem(0, QTableWidgetItem("Employee"))
        for i in range(1, num_days + 1):
            date = datetime.date(self.current_year, self.current_month, i)
            day_number = str(i)
            day_name = date.strftime("%A")  # Get the full name of the day
            header_text = f"{day_number}\n{day_name}"
            header_item = QTableWidgetItem(header_text)
            tableWidget.setItem(0, i, header_item)
            tableWidget.resizeRowsToContents()
        seen_keys = set()
        ordered_names = []

        for inner_dict in month_data.values():
            for key in inner_dict.keys():
                if key not in seen_keys:
                    ordered_names.append(key)
                    seen_keys.add(key)
        for row_index, employee_name in enumerate(ordered_names, start=1):
            tableWidget.setItem(row_index,0,QTableWidgetItem(employee_name.strip()))
        # Update the table with the retrieved data
        for day, inner_dict in month_data.items():
            for employee_name, (absence_days, absence_type) in inner_dict.items():
                employee_name = employee_name.strip()  # Trim whitespace from employee name
                # Find the row index corresponding to the employee name
                items = tableWidget.findItems(employee_name, Qt.MatchExactly)
                if items:
                    row_index = items[0].row()
                    # Calculate the column index based on the day
                    column_index = day
                    # Set the absence days in the table cell
                    for i in range(int(absence_days)):

                        if column_index + i <= num_days:  # Ensure it doesn't exceed the maximum day
                            actual_day=datetime.date(self.current_year,self.current_month,column_index+i)
                            pandas_day=pd.to_datetime(actual_day)
                            if pandas_day.dayofweek<5:
                                color, _ = self.get_absence_type_color(absence_type)
                                cell_item = QTableWidgetItem(self.get_absence_letter(absence_type))
                                cell_item.setBackground(color)
                                tableWidget.setItem(row_index, column_index+i, cell_item)
                            else:
                                cell_item = QTableWidgetItem('W')
                                light_grey = QColor(211,211,211)  # Adjust the RGB values for the desired shade of gray
                                cell_item.setBackground(light_grey)
                                tableWidget.setItem(row_index, column_index+i, cell_item)
                            if absence_days == 1.5:
                                cell_item = QTableWidgetItem('H1')
                                light_green = QColor(144, 238, 144)
                                cell_item.setBackground(light_green)
                                tableWidget.setItem(row_index, column_index+i, cell_item)
        ro_holidays = self.get_national_holidays(self.current_year,self.current_month)

        light_grey = QColor(211,211,211)  # Adjust the RGB values for the desired shade of gray

        for row_index, employee_name in enumerate(ordered_names, start=1):
            for i in range(1, num_days + 1):
                if QDate(self.current_year, self.current_month, i) in ro_holidays:
                    cell_item = QTableWidgetItem('B')
                    cell_item.setBackground(light_grey)
                    tableWidget.setItem(row_index, i, cell_item)

        for col in range(tableWidget.columnCount()):
            item = tableWidget.item(0, col)
            if item is None:
                item = QTableWidgetItem()
                tableWidget.setItem(0, col, item)
            item.setBackground(light_blue)
        for row_index in range(1, tableWidget.rowCount()):  # Iterate through rows
            total_days = 0.0
            for col_index in range(1, num_days + 1):  # Iterate through columns (days)
                item = tableWidget.item(row_index, col_index)
                if item is not None:
                    absence_days = item.text()  # Get absence days for the current cell
                    for day in absence_days:
                        print(day)
                        if  day == '1':
                            total_days -= 0.5
                        else:  
                            if day == 'H':
                                total_days += 1  
    
            # Add the total days taken to a new column at the end of the row
            total_days_item = QTableWidgetItem(str(total_days))
            tableWidget.setItem(row_index, num_days , total_days_item)
        tableWidget.setItem(0,num_days,QTableWidgetItem('Total'))
        tableWidget.resizeRowsToContents()
        tableWidget.resizeColumnsToContents()
        tableWidget.repaint()
        return tableWidget
    
class ApplicationWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.selections = {
            'department': None,
            'project': None,
            'employee': None,
            'leave': None,
            'manager': None,
            'period': None
        }
        self.title = 'Vacation Rate App'
        self.currentDialog = None  
        self.initUI()


    def openMonthlyTable(self):
        monthly_table_window = MonthlyTableWindow(self)
        monthly_table_window.exec_()


    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(100, 100, 1200, 760)  # Window size

        # Create Menu Bar
        menubar = self.menuBar()
        viewMenu = menubar.addMenu('View')
        downloadMenu = menubar.addMenu('Download')
        helpMenu = menubar.addMenu('Help')

        # Main layout
        mainLayout = QHBoxLayout()

        # Left panel for filter buttons
        leftPanel = QVBoxLayout()

        filtersGroup = QGroupBox("Filters")
        viewGroup = QGroupBox("View")
        employeesGroup = QGroupBox("Employees")

        # Create layouts for each GroupBox
        filtersLayout = QVBoxLayout()
        viewLayout = QVBoxLayout()
        employeesLayout = QVBoxLayout()



        # Setup filter buttons
        self.monthlyCalendarViewButton  = QPushButton('Monthly Calendar View')
        self.monthlyCalendarViewButton.clicked.connect(self.openMonthlyTable)
        #self.modifyButtonAppearance(self.monthlyCalendarViewButton )

        self.selectSpecificPeriodButton = QPushButton('Select Specific Period')
        self.selectSpecificPeriodButton.clicked.connect(self.showPeriodDialog)
        self.departmentButton = QPushButton('Department')
        self.departmentButton.clicked.connect(lambda: self.showSelectionDialog(unique_departments, 'Select Department'))
        
        self.managerButton = QPushButton('Manager')
        self.managerButton.clicked.connect(lambda: self.showSelectionDialog(unique_managers, 'Select Manager'))
        self.projectButton = QPushButton('Project')
        self.projectButton.clicked.connect(lambda: self.showSelectionDialog(unique_project, 'Select Project'))
        self.employeeButton = QPushButton('Employee')
        self.employeeButton.clicked.connect(lambda: self.showSelectionDialog(unique_employee, 'Select Employee'))
        self.typeOfLeaveButton = QPushButton('Type of Leave')
        self.typeOfLeaveButton.clicked.connect(lambda: self.showSelectionDialog(unique_leave, 'Select Type of Leave'))


        filtersLayout.addWidget(self.departmentButton)
        filtersLayout.addWidget(self.projectButton)
        filtersLayout.addWidget(self.managerButton)
        filtersLayout.addWidget(self.employeeButton)
        filtersLayout.addWidget(self.typeOfLeaveButton)

        viewLayout.addWidget(self.monthlyCalendarViewButton)
        viewLayout.addWidget(self.selectSpecificPeriodButton)
        self.outputTextEdit = QTextEdit()
        employeesLayout.addWidget(self.outputTextEdit)
        # Quarterly Buttons
        quarterlyLayout = QHBoxLayout()
        quarterlyLabel = QLabel("Quarterly")
        quarterlyLayout.addWidget(quarterlyLabel)
        for q in range(1, 5):
            quarterButton = QPushButton(f"Q{q}")
            quarterButton.setFixedWidth(50)  # Adjust width as needed
            # Connect to a method to handle the quarter selection
            quarterButton.clicked.connect(lambda checked, b=quarterButton: self.handlePredefinedPeriod(b.text()))
            quarterlyLayout.addWidget(quarterButton)
        viewLayout.addLayout(quarterlyLayout)

        # Semester Buttons
        semesterLayout = QHBoxLayout()
        semesterLabel = QLabel("Semester")
        semesterLayout.addWidget(semesterLabel)
        for s in range(1, 3):
            semesterButton = QPushButton(f"S{s}")
            semesterButton.setFixedWidth(50)  # Adjust width as needed
            # Connect to a method to handle the semester selection
            semesterButton.clicked.connect(lambda checked, b=semesterButton: self.handlePredefinedPeriod(b.text()))
            semesterLayout.addWidget(semesterButton)
        semesterLayout.addStretch(1)
        viewLayout.addLayout(semesterLayout)

        # Annual Button
        self.annualButton = QPushButton('Annual')

        viewLayout.addWidget(self.annualButton)
        self.annualButton.clicked.connect(lambda checked: self.handlePredefinedPeriod('All Available Periods'))

        viewGroup.setLayout(viewLayout)


        # Right panel for the graph and metrics, now including a tabbed interface
        rightPanel = QVBoxLayout()
        self.tabWidget = QTabWidget()
        self.histogramTab = QWidget()
        self.doughnutChartTab = QWidget()
        self.remainingLeavesTab = QWidget()

        self.tabWidget.addTab(self.histogramTab, "Histogram")
        self.tabWidget.addTab(self.doughnutChartTab, "Doughnut Chart")
        self.tabWidget.addTab(self.remainingLeavesTab, "Remaining Leaves")

        # Setup layouts for the histogram and the doughnut chart tabs
        self.histogramLayout = QVBoxLayout(self.histogramTab)
        self.doughnutChartLayout = QVBoxLayout(self.doughnutChartTab)
        self.remainingLeavesLayout = QVBoxLayout(self.remainingLeavesTab)

        # Setup the histogram tab with a matplotlib canvas
        self.figureHistogram = Figure()
        self.canvasHistogram = FigureCanvas(self.figureHistogram)
        self.histogramLayout.addWidget(self.canvasHistogram)

        # Doughnut Chart
        self.figureDoughnut = Figure(figsize=(10, 7))
        self.canvasDoughnut = FigureCanvas(self.figureDoughnut)
        self.doughnutChartLayout.addWidget(self.canvasDoughnut)

        # Remaining Leaves Chart
        self.figureLeaves = Figure(figsize=(10, 6))
        self.canvasLeaves = FigureCanvas(self.figureLeaves)
        self.remainingLeavesLayout.addWidget(self.canvasLeaves)

        # No need to setup the doughnut chart here; it will be initialized in createDoughnutChart()

        rightPanel.addWidget(self.tabWidget)

        filtersGroup.setLayout(filtersLayout)
        viewGroup.setLayout(viewLayout)
        employeesGroup.setLayout(employeesLayout)

        # Adjust the main layout to use the new group boxes
        leftPanel = QVBoxLayout()
        leftPanel.addWidget(filtersGroup)
        leftPanel.addWidget(viewGroup)
        leftPanel.addWidget(employeesGroup)

        # Combine left and right panels into the main layout
        mainLayout.addLayout(leftPanel, 1)
        mainLayout.addLayout(rightPanel, 4)

        # Set the central widget and show the main window
        centralWidget = QWidget()
        centralWidget.setLayout(mainLayout)
        self.setCentralWidget(centralWidget)
        self.createHistogram()
        
        self.createDoughnutChart()
        self.createRemainingLeavesChart()
        self.show()



    
    def modifyButtonAppearance(self, button):
        button.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        button.setFont(QFont("Arial", 14))  # Adjust the font and size as needed

        # Apply the filter_buttons_stylesheet to the button
        button.setStyleSheet(stylesheet)

        button.setCursor(Qt.PointingHandCursor)  # Change cursor to a hand when hovering


    def onDialogClosed(self):
     self.currentDialog = None

    
    def showPeriodDialog(self):
        if self.currentDialog is not None:
            self.currentDialog.close()
        self.currentDialog = PeriodDialog(self)
        self.currentDialog.customPeriodSelected.connect(self.handleCustomPeriod)  # Connect to a slot for custom period
        self.currentDialog.predefinedPeriodSelected.connect(self.handlePredefinedPeriod)  # Connect to a slot for predefined period
        self.currentDialog.show()
        self.currentDialog.finished.connect(self.onDialogClosed)

    def handleCustomPeriod(self, startDate, endDate):
        # Convert startDate and endDate from strings to datetime objects
        start_date = pd.to_datetime(startDate)
        end_date = pd.to_datetime(endDate)

        # Store the custom period as a tuple of (start_date, end_date)
        self.selections['period'] = (start_date, end_date)
        print(f"Custom period received in main window: {start_date} to {end_date}")
        print("Current selections:", self.selections)
        self.createHistogram() 
        self.createDoughnutChart() 


    def handlePredefinedPeriod(self, period):
        global df  

        # Calculate the most common year in the data
        year_counts = df['From'].dt.year.value_counts() + df['To'].dt.year.value_counts()
        if year_counts.empty:
            print("The dataset does not contain valid date information.")
            self.selections['period'] = None
            return
        selected_year = year_counts.idxmax()

        # Ensure 'From' and 'To' columns are datetime
        df['From'] = pd.to_datetime(df['From'])
        df['To'] = pd.to_datetime(df['To'])

        if period == "All Available Periods":
            self.selections['period'] = None
            print("All available data will be displayed.")
        elif period in ['Q1', 'Q2', 'Q3', 'Q4']:
            quarter = int(period[1])
            start_month = (quarter - 1) * 3 + 1
            end_month = start_month + 2
            start_date = pd.Timestamp(selected_year, start_month, 1)
            end_date = pd.Timestamp(selected_year, end_month, 1) + pd.offsets.MonthEnd()
            self.selections['period'] = (start_date, end_date)
            print(f"Period set to: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        elif period in ['S1', 'S2']:
            semester = int(period[1])
            start_date = pd.Timestamp(selected_year, 1 if semester == 1 else 7, 1)
            end_date = pd.Timestamp(selected_year, 6 if semester == 1 else 12, 30)
            self.selections['period'] = (start_date, end_date)
            print(f"Period set to: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        else:
            month_num = [QDate.longMonthName(i) for i in range(1, 13)].index(period) + 1
            start_date = pd.Timestamp(selected_year, month_num, 1)
            end_date = start_date + pd.offsets.MonthEnd()
            self.selections['period'] = (start_date, end_date)
            print(f"Period set to: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        
        self.createHistogram()
        self.createDoughnutChart()
        self.createRemainingLeavesChart()


    
    def handleSelection(self, selections, category):
        # category is one of 'Department', 'Project', 'Employee', or 'Type of Leave'
        category_key_map = {
            'Select Department': 'department',
            'Select Project': 'project',
            'Select Employee': 'employee',
            'Select Type of Leave': 'leave',
            'Select Manager' : 'manager'
        }
        category_key = category_key_map.get(category)
        if category_key:
            self.selections[category_key] = selections
            print(f"{category_key} selected: {selections}")
        print("Current selections:", self.selections)
        self.createHistogram()
        self.createDoughnutChart()
        self.createRemainingLeavesChart()



    def showSelectionDialog(self, options, title):
        # Convert options to strings and filter out NaN values
        options = [str(option) for option in options if pd.notnull(option)]
        if self.currentDialog is not None:
            self.currentDialog.close()
        self.currentDialog = SelectionDialog(options, title, self)
        self.currentDialog.selectionMade.connect(self.handleSelection)  # Connect the signal to handleSelection
        self.currentDialog.show()
        self.currentDialog.finished.connect(self.onDialogClosed)

    def resizeEvent(self, event):
        QMainWindow.resizeEvent(self, event)
        # Reposition the metrics frame when the main window is resized

    def calculate_remaining_leaves(self):
        df = self.filterData(without_period=False)
  
        annual_leave_df = df[df['Absence Type'] == 'Annual leave'].drop_duplicates(subset=['Employee ID', 'From', 'To', 'Absence Type'])
        leave_taken = annual_leave_df.groupby('Employee ID')['Att./abs. days'].sum()
        leave_entitlement = 25
        remaining_leaves = leave_entitlement - leave_taken
        
        # Dynamically determine the start of bins based on the data
        min_remaining_leave = min(0, remaining_leaves.min())  # Ensuring 0 is the minimum if negative values exist
        bins = np.arange(min_remaining_leave, 26.5)  # Adjust to ensure bins 0 to 25
        
        histogram_data, _ = np.histogram(remaining_leaves, bins=bins)
        return histogram_data, bins, remaining_leaves


    def determine_bin_size(self):
        """Determine the bin size ('day', 'week', 'month') based on the selected period."""
        if self.selections['period']:
            start_date, end_date = self.selections['period']
            delta = end_date - start_date
            # Use days for comparison, avoiding ambiguous 'M' or 'Y' units
            if delta.days <= 30:  # Approximating 1 month as 30 days
                return 'day'
            elif 30 < delta.days <= 90:  # Approximating 3 months as 90 days
                return 'week'
            else:
                return 'month'
        return 'month'  # Default bin size
    def total_working_days(self,start_date,end_date):
        current_date = pd.to_datetime(end_date)
        total_days = 0
        # Iterate over each day from start_date to current_date
        current=pd.to_datetime(start_date)
        while current <= current_date:
            # Check if the current day is a weekday (Monday=0, Sunday=6)
            if current.weekday() < 5:
                total_days += 1
            current += datetime.timedelta(days=1)  # Move to the next day
        return total_days
    def calculate_duration_excluding_weekends(self,start_date, end_date):
        weekdays = pd.date_range(start=start_date, end=end_date, freq='B')  # Generate weekdays between start and end dates
        return len(weekdays)
    def calculate_metrics(self):
        filtered_data=self.selections
        
        metrics_df=df
        all_days=self.total_working_days(metrics_df['From'].min(),metrics_df['To'].max())
        if filtered_data['employee'] is not None:
            metrics_df=df[df['Employee Name'].isin(filtered_data['employee'])]
            metrics_df=metrics_df.drop_duplicates(subset=['Employee Name', 'From'])
        else: 
            if filtered_data['project'] is not None:
                metrics_df=df[df['Project Name'].isin(filtered_data['project'])]

            else:
                if filtered_data['department'] is not None:
                    metrics_df = df[df['Departament'].isin(filtered_data['department'])]
                    metrics_df=metrics_df.drop_duplicates(subset=['Employee Name', 'From'])
                else:
                    if filtered_data['manager'] is not None:
                        metrics_df = df[df['Manager Name'].isin(filtered_data['manager'])]
                        metrics_df=metrics_df.drop_duplicates(subset=['Employee Name', 'From'])
        if filtered_data['period']:
            start_date, end_date = self.selections['period']
            all_days=self.total_working_days(start_date,end_date)
            metrics_df['From'] = pd.to_datetime(metrics_df['From'])
            metrics_df['To'] = pd.to_datetime(metrics_df['To'])
            metrics_df = metrics_df[
                    ((metrics_df['To'] >= start_date))]
            metrics_df = metrics_df[
                    ((metrics_df['From'] <=end_date))]
        metrics_df['From'] = pd.to_datetime(metrics_df['From'])
        metrics_df['To'] = pd.to_datetime(metrics_df['To'])
        unique_employees=metrics_df["Employee Name"].unique()
        total_absence_days_with_req=metrics_df[metrics_df['Request Status']!='Rejected']
        total_absence_days_with_req=total_absence_days_with_req[~total_absence_days_with_req["Project Name"].isin(['Intercontract','Newcomer','No contract'])]
        total_absence_days=total_absence_days_with_req['Att./abs. days'].sum()
        total_employees=len(unique_employees)
        if filtered_data['period']:
            start_result = start_date > metrics_df['From']
            end_result = end_date< metrics_df['To']
            if start_result.any():
                metrics_df.loc[start_result, 'From'] = start_date
            if end_result.any():
                metrics_df.loc[end_result, 'To'] = end_date
        # Calcuate duration of each absence
        metrics_df['Duration'] = metrics_df.apply(lambda row: self.calculate_duration_excluding_weekends(row['From'], row['To']), axis=1)

        total_days_possible=all_days*total_employees
        # Filter the DataFrame based on project name
        filtered_df = metrics_df[metrics_df['Project Name'].isin(['Intercontract', 'Newcomer', 'No project'])]
        # Compute total days taken for each group
        total_days_taken = filtered_df.groupby('Project Name')['Duration'].sum().reset_index()
        total_days_taken.rename(columns={'Duration': 'Total Days Taken'}, inplace=True)
        percentage_productive=str(((total_days_possible-total_days_taken['Total Days Taken'].sum()-total_absence_days)/total_days_possible)*100)
        appended_value_productivity=percentage_productive[:5]
        appended_productivity=f"Productivity rate for selected filters is {appended_value_productivity}%"
        self.outputTextEdit.clear()
        self.outputTextEdit.append(appended_productivity)
        appended_value_co=str(total_absence_days/total_days_possible*100)
        appended_text_co=f"Days off rate for selected filters is {appended_value_co[:4]}%"
        metrics_df_without_duplicates = metrics_df.drop_duplicates(subset=["Employee ID"])
        entitlement_days=metrics_df_without_duplicates["Sum of Entitlement"].sum()
        total_annual_leave_df=metrics_df[metrics_df['Absence Type']=='Annual leave']
        total_annual_leave_df=total_annual_leave_df[total_annual_leave_df['Request Status']=='Approved']
        total_annual_leave=total_annual_leave_df['Att./abs. days'].sum()
        appended_value_remaining=str(entitlement_days-total_annual_leave)
        appended_text_remaining=f"Days off remaining number: {appended_value_remaining}"
        self.outputTextEdit.append(appended_text_co)
        self.outputTextEdit.append(appended_text_remaining)
        self.outputTextEdit.append(f"Number of employees: {len(unique_employees)}")
        self.outputTextEdit.append('\n')
        self.outputTextEdit.append(f"Information valid for")
        for i in self.selections:
            self.outputTextEdit.append(f"{i}:{self.selections[i]}")
        
    def aggregate_data(self, bin_size):
        filtered_df = self.filterData(without_period=False)
        print(filtered_df)
        if filtered_df.empty:
            return pd.DataFrame()

        filtered_df['From'] = pd.to_datetime(filtered_df['From'])
        filtered_df['To'] = pd.to_datetime(filtered_df['To'])
        filtered_df_valid=filtered_df[filtered_df['Request Status']!='Rejected']
        if self.selections.get('project') is None:
            filtered_df_valid = filtered_df_valid.drop_duplicates(subset=['Employee ID', 'From', 'To', 'Absence Type'])

        # Initialize an empty list to store all absences
        all_absences = []

        # Loop through the filtered DataFrame
        for _, row in filtered_df_valid.iterrows():
            business_days = pd.date_range(row['From'], row['To']).to_series().map(lambda x: x if x.weekday() < 5 else None).dropna()
            
            # Calculate the abs_days_per_date for distributing 'Att./abs. days' evenly across business days
            abs_days_per_date = row['Att./abs. days'] / len(business_days) if len(business_days) > 0 else row['Att./abs. days']
            
            # Special handling for single business day with half-day absence
            if len(business_days) == 1 and row['Att./abs. days'] == 0.5:
                all_absences.append({'Date': business_days.iloc[0], 'AbsenceDays': 0.5})
            else:
                for date in business_days:
                    all_absences.append({'Date': date, 'AbsenceDays': abs_days_per_date})
        if filtered_df_valid.empty:
            return pd.DataFrame
        all_absences_df = pd.DataFrame(all_absences)
        # Aggregate the absences based on the bin size (day, week, or month)
        aggregated_all_df = self.aggregate_absences(all_absences_df, bin_size)

        filtered_annual_leave_df = self.filterData(without_period=True)
        annual_leave_absences = []
        filtered_annual_leave_df_valid=filtered_annual_leave_df[filtered_annual_leave_df['Request Status']!='Rejected']
        for _, row in filtered_annual_leave_df_valid.iterrows():
            business_days = pd.date_range(row['From'], row['To']).to_series().map(lambda x: x if x.weekday() < 5 else None).dropna()
            for date in business_days:
                annual_leave_absences.append({'Date': date, 'AbsenceDays': 1})
        annual_leave_absences_df = pd.DataFrame(annual_leave_absences)
        if self.selections['leave'] is None or 'Annual leave' in self.selections['leave']:
            aggregated_annual_leave_df = self.aggregate_absences(annual_leave_absences_df, bin_size)
            aggregated_annual_leave_df['CumulativeAbsenceDays'] = aggregated_annual_leave_df['AbsenceDays'].cumsum()
            unique_entitlements = filtered_df.drop_duplicates(subset=['Employee Name'])
            total_entitlement = unique_entitlements['Sum of Entitlement'].sum()
            aggregated_annual_leave_df['CumulativePercentage'] = (aggregated_annual_leave_df['CumulativeAbsenceDays'] / total_entitlement) * 100
            # Combine the aggregated dataframes
            aggregated_df = aggregated_all_df.merge(aggregated_annual_leave_df[['Date', 'CumulativePercentage']], on='Date', how='left').fillna(method='ffill')
        else:
            # If "Annual leave" is not among the selected leave types, skip the cumulative percentage calculation
            aggregated_df = aggregated_all_df.copy()
            aggregated_df['CumulativePercentage'] = None


        return aggregated_df



    def aggregate_absences(self, absences_df, bin_size):
        """Aggregate absence days based on the bin_size."""
        absences_df['Date'] = pd.to_datetime(absences_df['Date'])
        if bin_size == 'day':
            aggregated_df = absences_df.groupby(absences_df['Date'].dt.date)['AbsenceDays'].sum().reset_index(name='AbsenceDays')
        elif bin_size == 'week':
            aggregated_df = absences_df.groupby(absences_df['Date'].dt.to_period('W'))['AbsenceDays'].sum().reset_index(name='AbsenceDays')
            aggregated_df['Date'] = aggregated_df['Date'].apply(lambda x: x.start_time.date())
        elif bin_size == 'month':
            aggregated_df = absences_df.groupby(absences_df['Date'].dt.to_period('M'))['AbsenceDays'].sum().reset_index(name='AbsenceDays')
            aggregated_df['Date'] = aggregated_df['Date'].apply(lambda x: x.start_time.date())
        return aggregated_df


    def createDoughnutChart(self):
        
        filtered_df = self.filterData()

        # Aggregate the leave types
        leave_counts = filtered_df.groupby('Absence Type')['Att./abs. days'].sum()
        total = leave_counts.sum()

        # Determine slices under 5%
        main_categories = leave_counts[leave_counts / total >= 0.05]
        others = leave_counts[leave_counts / total < 0.05].sum()

        # Add "Others" to main categories if necessary
        if others > 0:
            main_categories["Others"] = others

        # Create the main pie chart
        self.figureDoughnut.clear()  # Clear previous plot
        ax = self.figureDoughnut.add_subplot(111)
        wedges, texts, autotexts = ax.pie(main_categories.values, labels=main_categories.index, autopct='%1.1f%%', startangle=90)

        # Add a legend outside the pie chart
        ax.legend(wedges, main_categories.index, title="Leave Types", loc="upper right", bbox_to_anchor=(1.2, 1), fontsize='large', title_fontsize='large')

        ax.set_title('Leave Type Distribution', fontsize=18, color='black', fontweight='bold')
        plt.setp(texts, size='x-large', color='black', fontweight='bold')
        plt.setp(autotexts, size='x-large', color='white', weight='bold')

        # Handling "Others" slice details
        if others > 0:
            small_categories = leave_counts[leave_counts / total < 0.05]
            # Create a detailed text annotation for "Others"
            detailed_text = "Details for 'Others':\n" + "\n".join([f"{k}: {v/total*100:.1f}%" for k, v in small_categories.items()])
            ax.text(1, -1, detailed_text, ha='left', va='bottom', fontsize=10, fontweight='bold', bbox=dict(facecolor='white', alpha=0.5))

        self.figureDoughnut.tight_layout()

 
        self.canvasDoughnut.draw()




    def createHistogram(self):
        bin_size = self.determine_bin_size()
        aggregated_data = self.aggregate_data(bin_size)
        # Clear the previous figure
        self.figureHistogram.clear()
        ax = self.figureHistogram.add_subplot(111)
        ax2 = ax.twinx()  # Create a secondary y-axis for cumulative percentages


        if not aggregated_data.empty:

            if self.selections['period']:
                start_date, end_date = self.selections['period']
                if bin_size == 'day':
                    # For daily bins, use every day in the period
                    complete_date_range = pd.date_range(start=start_date, end=end_date).to_frame(index=False, name='Date')
                elif bin_size == 'week':
                    # For weekly bins, start each bin on the first day of the week in the period
                    complete_date_range = pd.date_range(start=start_date, end=end_date, freq='W-MON').to_frame(index=False, name='Date')
                elif bin_size == 'month':
                    # For monthly bins, start each bin on the first day of the month in the period
                    complete_date_range = pd.date_range(start=start_date, end=end_date, freq='MS').to_frame(index=False, name='Date')

                # Ensure aggregated_data 'Date' is datetime for merging
                aggregated_data['Date'] = pd.to_datetime(aggregated_data['Date'])

                # Merge to ensure all bins are included
                aggregated_data = pd.merge(complete_date_range, aggregated_data, on='Date', how='left')

                aggregated_data['AbsenceDays'] = aggregated_data['AbsenceDays'].fillna(0)
                aggregated_data['CumulativePercentage'] = aggregated_data['CumulativePercentage'].fillna(method='ffill')
                aggregated_data['Date'] = aggregated_data['Date'].dt.date
                
            # Dynamically adjust the bar width based on bin size
            bar_width = {'day': 0.7, 'week': 5, 'month': 20}.get(bin_size, 0.7)

            # Plot the histogram with the specified color and add a label for the blue bars
            bars = ax.bar(aggregated_data['Date'] , aggregated_data['AbsenceDays'], width=bar_width, color=(173/255, 216/255, 230/255), alpha=0.7, label='Absence Days Taken')

            # Plotting the cumulative percentage and adding a label for the red line
            ax2.plot(aggregated_data['Date'] , aggregated_data['CumulativePercentage'], color='red', marker='o', linestyle='-', label='Cumulative Days Taken (%)')
            ax2.set_ylabel('Cumulative Days Taken (%)', color='red')
            ax2.tick_params(axis='y', colors='red')
            
            # Adjust x-axis formatting based on bin size
            if bin_size == 'day':
                ax.xaxis.set_major_locator(mdates.DayLocator())
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m-%Y'))
            elif bin_size == 'week':
                ax.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=mdates.MONDAY))
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%d%b'))
                
                # Generate labels for the start and end of each week
                week_labels = [
                    '{0:%d%b}-{1:%d%b}'.format(start, start + pd.Timedelta(days=6))
                    for start in pd.date_range(start_date, end_date, freq='W-MON')
                ]
                
                # Set custom tick labels
                ax.set_xticklabels(week_labels)
            elif bin_size == 'month':
                # For monthly bins, show the abbreviated month name and year
                ax.xaxis.set_major_locator(mdates.MonthLocator())
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
            ax.figure.autofmt_xdate()

            # Annotate each bin with its value
            for bar in bars:
                height = bar.get_height()
                ax.annotate(f'{int(height)}',
                            xy=(bar.get_x() + bar.get_width() / 2, height),
                            xytext=(0, 3),  # 3 points vertical offset
                            textcoords="offset points",
                            ha='center', va='bottom')

            # Add a grid for better readability
            ax.grid(True, which='both', linestyle='--', linewidth=0.5, color='grey', alpha=0.5)

            # Set titles and labels
            ax.set_title('Absence Counts')
            ax.set_xlabel('Period')
            ax.set_ylabel('Total Absence Days')

            # Combining legends from both axes
            handles, labels = ax.get_legend_handles_labels()
            handles2, labels2 = ax2.get_legend_handles_labels()
            ax2.legend(handles + handles2, labels + labels2, loc='upper left')

        else:
            # Display message if no data
            ax.text(0.5, 0.5, 'No data to display for the selected filters', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)

        self.figureHistogram.subplots_adjust(left=0.07, right=0.95, top=0.95, bottom=0.15)
        self.canvasHistogram.draw()
        self.calculate_metrics()


    def createRemainingLeavesChart(self):
        histogram_data, bins, remaining_leaves = self.calculate_remaining_leaves()
        
        self.figureLeaves.clear()
        ax = self.figureLeaves.add_subplot(111)
        
        bin_centers = 0.5 * (bins[:-1] + bins[1:])
        mean = np.mean(remaining_leaves)
        std_dev = np.std(remaining_leaves, ddof=1)
        
        ax.bar(bin_centers, histogram_data, align='center', color='skyblue', label='Number of Employees')
        
        if std_dev > 0 and not np.isnan(std_dev):
            x = np.linspace(mean - 3*std_dev, mean + 3*std_dev, 100)
            y = norm.pdf(x, mean, std_dev)
            
            # Scale y by total counts to make it comparable to histogram
            scaled_y = y * sum(histogram_data) * np.diff(bins[:2])  # np.diff(bins[:2]) gives bin width
            
            ax2 = ax.twinx()
            percentage_y = scaled_y / sum(histogram_data) * 100  # Convert to percentages of total counts
            ax2.plot(x, percentage_y, 'r-', label='Normal Distribution (%)')
            ax2.set_ylabel('Percentage (%)')
            ax2.legend(loc='upper right')
        else:
            print("Standard deviation is zero or NaN. Normal distribution curve will not be plotted.")
        
        ax.set_xlabel('Remaining Annual Leave Days')
        ax.set_ylabel('Number of Employees')
        ax.set_title('Distribution of Remaining Annual Leave Days with Normal Distribution')
        ax.legend(loc='upper left')
        
        ax.set_xticks(bin_centers)
        ax.set_xticklabels([f"{int(bin)}" for bin in bins[:-1]])
        
        ax.grid(True, which='both', linestyle='--', linewidth=0.5, color='grey', alpha=0.5)
        self.figureLeaves.tight_layout() 
        
        self.canvasLeaves.draw()


    def displayMessage(self, message):
        # Clear the previous figure and display a message
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.text(0.5, 0.5, message, horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
        self.canvas.draw()

    
    def filterData(self, without_period=False):
        global df
        filtered_df = df.copy()

        if without_period:
            # Filter for "Annual leave" and apply other selections except for 'leave'
            filtered_df = filtered_df[filtered_df['Absence Type'] == 'Annual leave']
            for category, selection in self.selections.items():
                if selection and category not in ['leave']:  # Skip 'leave' type filtering
                    column_map = {
                        'department': 'Departament',
                        'project': 'Project Name',
                        'employee': 'Employee Name',
                        'manager':'Manager Name',
                    }
                    if category in column_map:
                        filtered_column = column_map[category]
                        filtered_df = filtered_df[filtered_df[filtered_column].isin(selection)]
        else:

            # Apply period filter
            if self.selections['period']:
                start_date, end_date = self.selections['period']
                filtered_df = filtered_df[
                    ((filtered_df['From'] >= start_date) & (filtered_df['To'] <= end_date)) |
                    ((filtered_df['From'] <= end_date) & (filtered_df['To'] >= start_date))
                ]
            # Apply filters for all categories
            for category, selection in self.selections.items():
                if selection:
                    column_map = {
                        'department': 'Departament',
                        'project': 'Project Name',
                        'employee': 'Employee Name',
                        'leave': 'Absence Type',
                        'manager': 'Manager Name',
                    }
                    if category in column_map:
                        filtered_column =  column_map[category]
                        filtered_df = filtered_df[filtered_df[filtered_column].isin(selection)]

        return filtered_df




# Run the application
app = QApplication(sys.argv)


app.setStyleSheet(stylesheet)
ex = ApplicationWindow()
sys.exit(app.exec_())
